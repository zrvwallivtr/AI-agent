import os
from os.path import exists
import re
import json
import pdfplumber
import openpyxl
import docx
import ebooklib
import mimetypes
import warnings
from typing import Optional
from pathlib import Path
from ebooklib import epub
from bs4 import BeautifulSoup
from typing import Any
from docling.document_converter import DocumentConverter

from src import config
from src.logger import get_logger


logger = get_logger(__name__)


class PathTraversalError(Exception):
    pass


UPLOAD_DIR = Path(f"~/.{config.APP_NAME}/uploads").expanduser().resolve()

def _validate_path(path: Path) -> Path:
    """Resolve 'path' to its real, absolute form."""
    resolved = path.resolve()

    try:
        resolved.relative_to(UPLOAD_DIR)

    except ValueError:
        raise PathTraversalError(
            f"Path '{path}' resolves to '{resolved}', which is outside "
            f"the allowed directory '{UPLOAD_DIR}'"
        )

    return resolved


class Parsers:
    def __init__(
        self,
    ):
        # Map formats to their parsing methods
        self.formats = {
            # Plain text
            ".txt": self.read_txt,

            # Data & configuration formats
            ".csv": self._docling_read_csv if config.DOCLING_DEFAULT else self._read_csv,
            ".xlsx": self._read_xlsx,
            ".yaml": self._read_yaml,
            ".yml": self._read_yml,
            ".toml": self._read_toml,
            ".xml": self._read_xml,
            
            # Text documents
            ".pdf": self._docling_read_pdf if config.DOCLING_DEFAULT else self._read_pdf,
            ".docx": self._read_docx,
            ".epub": self._docling_read_epub if config.DOCLING_DEFAULT else self._read_epub,
            
            # Programming
            ".py": self._read_code,
            ".js": self._read_code,
            ".ts": self._read_code,
            ".tsx": self._read_code,
            ".json": self._read_code,
            ".md": self._read_code,
            ".sh": self._read_code,
            ".html": self._read_code,
            ".css": self._read_code,
            ".rs": self._read_code,
            ".go": self._read_code,
        }

        self.converter = DocumentConverter()

    # =======================================================
    # Data & configuration formats
    # =======================================================

    def _read_csv(self, path: Path) -> str:
        """Reads csv as plain text files."""
        return self.read_txt(path, "csv")

    def _docling_read_csv(self, path: Path) -> str:
        """Convert csv to markdown via docling."""
        logger.info(f"Converting CSV file: {path}")
        print(f"Converting CSV file: {path}")
        result = self.converter.convert(path)
        return result.document.export_to_markdown()

    def _read_xlsx(self, path: Path) -> str:
        """Extracts text content from spreadsheet row by row."""
        try:
            wb          = openpyxl.load_workbook(path, data_only=True)
            excel_text  = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                excel_text.append(f"=== Sheet: {sheet_name} ===")

                for row in ws.iter_rows(values_only=True):
                    # Only process lines that contain actual data elements
                    if any(cell is not None for cell in row):
                        clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
                        excel_text.append(" | ".join(clean_row))

            logger.info("Extracted text content from xlsx file row by row: path='%s', sheets=%d", path, len(wb.sheetnames))
            return "\n".join(excel_text)

        except Exception as e:
            logger.error("Failed to read xlsx file: path=%s, error=%s", path, e)
            return f"Failed to extract file content"

    def _read_yaml(self, path: Path) -> str:
        return self.read_txt(path, "yaml")

    def _read_yml(self, path: Path) -> str:
        return self.read_txt(path, "yml")

    def _read_toml(self, path: Path) -> str:
        return self.read_txt(path, "toml")

    def _read_xml(self, path: Path) -> str:
        return self.read_txt(path, "xml")

    # =======================================================
    # Text documents
    # =======================================================

    def _read_pdf(self, path: Path) -> str:
        """Extracts pdf text content layout page by page."""
        try:
            with pdfplumber.open(path) as pdf:
                pages = [page.extract_text() for page in pdf.pages]
                pages = [p for p in pages if p]
                logger.info("Extracted PDF text content page by page: path='%s', pages=%d", path, len(pdf.pages))
                return "\n\n".join(pages)

        except Exception as e:
            logger.error("Failed to read PDF file: path=%s, error=%s", path, e)
            return f"Failed to extract file content"

    def _docling_read_pdf(self, path: Path) -> str:
        """Convert pdf to markdown via docling."""
        logger.info(f"Converting PDF file: {path}")
        print(f"Converting PDF file: {path}")
        result = self.converter.convert(path)
        return result.document.export_to_markdown()
    
    def _read_docx(self, path: Path) -> str:
        """Extracts structural text elements line by line from document."""
        try:
            doc = docx.Document(path)
            paragraphs = []

            for para in doc.paragraphs:
                clean_text = para.text.strip()
                if clean_text:
                    paragraphs.append(clean_text)
            
            logger.info("Extracted docx text content line by line: path='%s', paragraphs=%d", path, len(paragraphs))
            return "\n".join(paragraphs)

        except Exception as e:
            logger.error("Failed to read DOCX file: path='%s', error=%s", path, e)
            return f"Failed to extract file content"

    def _read_epub(self, path: Path) -> str:
        """Extracts plain text blocks from internal EPUB XHTML document payloads."""
        try:
            book            = epub.read_epub(path)
            chapters_text   = []

            for item in book.get_items():
                # EPUB structural content is split into internal document items
                if item.get_type() == ebooklib.ITEM_DOCUMENT:
                    soup = BeautifulSoup(item.get_content(), "lxml")

                    # Remove visual layout nodes that shouldn't feed context windows
                    for junk in soup(["script", "style"]):
                        junk.decompose()

                    plain_text = soup.get_text(separator="\n").strip()
                    if plain_text:
                        chapters_text.append(plain_text)

            logger.info(
                "Extracted plain text blocks from EPUB file: path='%s', chapters=%d",
                path,
                len(chapters_text)
            )
            return "\n\n".join(chapters_text)

        except Exception as e:
            logger.error("Failed to read EPUB file: path='%s', error=%s", path, e)
            return f"Failed to extract file content"

    def _docling_read_epub(self, path: Path) -> str:
        """Convert EPUB to markdown via docling."""
        logger.info(f"Converting EPUB file: {path}")
        print(f"Converting EPUB file: {path}")
        result = self.converter.convert(path)
        return result.document.export_to_markdown()

    # =======================================================
    # PROGRAMMING
    # =======================================================

    def _read_code(self, path: Path) -> str:
        """Reads code files and wraps them in markdown code fences."""
        lang    = path.suffix.lstrip(".")
        content = self.read_txt(path, f"{lang}")

        return f"```{lang}\n{content}\n```"

    # =======================================================
    # Plain text
    # =======================================================

    def read_txt(self, path: Path, file_type: str = "txt") -> str:
        """Reads plain text files using UTF-8 validation."""
        try:
            cont = path.read_text(encoding="utf-8", errors="ignore")
            clean_cont = "".join(c for c in cont if c.isprintable() or c in "\n\r\t")
            logger.info("Extracted '%s' file content as plain text: path='%s'", file_type, path)
            return clean_cont
        except Exception as e:
            logger.error(f"Failed to read '%s' file: path='%s', error=%s", file_type, path, e)
            return f"Failed to extract file content"

    # =======================================================
    # READ
    # =======================================================

    def read_document(self, path: Path) -> str:
        """Return file content as a string using mapped parsers."""
        safe_path = _validate_path(path)

        if not safe_path.exists():
            raise FileNotFoundError(f"'{safe_path}' does not exist")
        if not safe_path.is_file():
            raise ValueError(f"'{safe_path}' is not a regular file")

        ext = safe_path.suffix.lower()
        parser = self.formats.get(ext, self.read_txt)
        return parser(safe_path)
