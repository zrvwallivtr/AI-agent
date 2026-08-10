import os
from os.path import exists
import re
import json
import pdfplumber
import openpyxl
import docx
import ebooklib
import mimetypes
import warnings
from typing import Optional
from pathlib import Path
from ebooklib import epub
from bs4 import BeautifulSoup
from typing import Any

from src.agent.llm import LLM
from src.agent.chat import Chat
from src import config


def session_path(session: str | None = None) -> Path:
    if session is not None:
        return config.DROPBOX_DIR / session
    return config.DROPBOX_DIR / "chat"

def _is_dir_empty(path: Path) -> bool:
    """Return True if directory is empty, otherwise 'False'."""
    if not path.exists():
        return False
    with os.scandir(path) as entries:
        return next(entries, None) is None

def _read_file(path) -> str:
    """Return content in file path."""
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        content = f.read()
    return content


class FileReader:
    def __init__(
        self,
        session: str | None = None,
    ):
        self.chat               = Chat(session=session)
        self.model              = config.MODEL

        # Files
        self.dropbox_dir        = session_path(session)
        self.file_or_not_prompt = config.FILE_OR_NOT_PROMPT
        self.file_list_prompt   = config.GET_FILE_LIST_PROMPT
        
        # Dropbox
        self.metadata_path      = session_path(session) / "file_metadata.json"
        self.file_metadata      = self._load_file_metadata()
        self.gen_summary_prompt = config.GEN_SUMMARY_PROMPT

        # Map formats to their parsing methods
        self.formats = {
            # Plain text
            ".txt": self._read_txt,

            # Data & configuration formats
            ".csv": self._read_csv,
            ".xlsx": self._read_xlsx,
            ".yaml": self._read_yaml,
            ".yml": self._read_yml,
            ".toml": self._read_toml,
            ".xml": self._read_xml,
            
            # Text documents
            ".pdf": self._read_pdf,
            ".docx": self._read_docx,
            ".epub": self._read_epub,
            
            # Programming
            ".py": self._read_code,
            ".js": self._read_code,
            ".ts": self._read_code,
            ".tsx": self._read_code,
            ".json": self._read_code,
            ".md": self._read_code,
            ".sh": self._read_code,
            ".html": self._read_code,
            ".css": self._read_code,
            ".rs": self._read_code,
            ".go": self._read_code,
        }

    # ==================================================
    # Parsers
    #
    # - Extracts file contents line by line into plain
    #   text.
    # ==================================================

    def _read_txt(self, path: Path, file_type: str = "txt") -> str:
        """Reads plain text files using UTF-8 validation."""
        try:
            return path.read_text(encoding="utf-8", errors="ignore")
        except Exception as e:
            return f"Error reading {file_type} file '{path}': {str(e)}"

    def _read_csv(self, path: Path) -> str:
        """Reads csv as plain text files."""
        return self._read_txt(path, "csv")

    def _read_xlsx(self, path: Path) -> str:
        """Extracts text content from spreadsheet row by row."""
        try:
            wb          = openpyxl.load_workbook(path, data_only=True)
            excel_text  = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                excel_text.append(f"=== Sheet: {sheet_name} ===")

                for row in ws.iter_rows(values_only=True):
                    # Only process lines that contain actual data elements
                    if any(cell is not None for cell in row):
                        clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
                        excel_text.append(" | ".join(clean_row))

            return "\n".join(excel_text)

        except Exception as e:
            return f"Error reading xlsx file '{path}': {str(e)}"

    def _read_yaml(self, path: Path) -> str:
        return self._read_txt(path, "yaml")

    def _read_yml(self, path: Path) -> str:
        return self._read_txt(path, "yml")

    def _read_toml(self, path: Path) -> str:
        return self._read_txt(path, "toml")

    def _read_xml(self, path: Path) -> str:
        return self._read_txt(path, "xml")

    def _read_pdf(self, path: Path) -> str:
        """Extracts pdf text content layout page by page."""
        try:
            with pdfplumber.open(path) as pdf:
                pages = [page.extract_text() for page in pdf.pages]
                pages = [p for p in pages if p]
                return "\n\n".join(pages)

        except Exception as e:
            return f"Error reading PDF file '{path}': {str(e)}"
    
    def _read_docx(self, path: Path) -> str:
        """Extracts structural text elements line by line from document."""
        try:
            doc = docx.Document(path)
            paragraphs = []

            for para in doc.paragraphs:
                clean_text = para.text.strip()
                if clean_text:
                    paragraphs.append(clean_text)

            return "\n".join(paragraphs)

        except Exception as e:
            return f"Error reading docx file '{path}': {str(e)}"

    def _read_epub(self, path: Path) -> str:
        """Extracts plain text blocks from internal EPUB XHTML document payloads."""
        try:
            book            = epub.read_epub(path)
            chapters_text   = []

            for item in book.get_items():
                # EPUB structural content is split into internal document items
                if item.get_type() == ebooklib.ITEM_DOCUMENT:
                    soup = BeautifulSoup(item.get_content(), "lxml")

                    # Remove visual layout nodes that shouldn't feed context windows
                    for junk in soup(["script", "style"]):
                        junk.decompose()

                    plain_text = soup.get_text(separator="\n").strip()
                    if plain_text:
                        chapters_text.append(plain_text)

            return "\n\n".join(chapters_text)
        except Exception as e:
            return f"Error reading EPUB file '{path}': {str(e)}"

    def _read_code(self, path: Path) -> str:
        """Reads code files and wraps them in markdown code fences."""
        lang    = path.suffix.lstrip(".")
        content = self._read_txt(path, f"{lang}")

        return f"```{lang}\n{content}\n```"

    # ================================================
    # Metadata (file_metadata.json)
    # ================================================

    def _load_file_metadata(self) -> dict[str, dict[str, Any]]:
        """Fetch/create metadata file, returns metadata as dict."""
        if not self.dropbox_dir.exists():
            self.dropbox_dir.mkdir(parents=True, exist_ok=True)

        path = self.dropbox_dir / "file_metadata.json"

        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                return {}

        else:
            default_data = {}
            path.write_text(json.dumps(default_data, indent=4), encoding="utf-8")
            return default_data

    def _add_file_metadata(self, filename: str, summary: str):
        """
        Add metadata to 'file_metadata.json'.

        {
            "filename": {
                "summary": ,
                "path": ,
                "mime_type": ,
                "size_bytes":
            }
        }
        """
        path = self.dropbox_dir / filename
        mime_type, _ = mimetypes.guess_type(path)

        # Ensure 'file_metadata' is a dictionary
        if not isinstance(self.file_metadata, dict):
            raise TypeError(
                f"Error: Expected 'file_metadata' to be a dict, but got {type(self.file_metadata).__name__}"
                f"with value: {self.file_metadata!r}"
            )

        self.file_metadata[path.name] = {
            "summary": f"{summary}",
            "path": str(path),
            "mime_type": mime_type,
            "size_bytes": path.stat().st_size if path.exists() else 0
        }

        self.dropbox_dir.mkdir(parents=True, exist_ok=True)
        self.metadata_path.write_text(json.dumps(self.file_metadata, indent=4), encoding="utf-8")

    def _get_filenames_from_metadata(self) -> list[str]:
        """Return all filenames from 'file_metadata.json'."""
        if self.metadata_path.exists():
            try:
                data = json.loads(self.metadata_path.read_text(encoding="utf-8"))
                return list(data.keys())

            except json.JSONDecodeError:
                return []

        return []

    def _files_not_in_file_metadata(self) -> list[str]:
        """
        Return a list of available files that is
        not in 'file_data.json' (without summary).
        """
        files_with_summary = self._get_filenames_from_metadata()
        available_files = self.list_available_files()

        files_without_summary = list(set(available_files) - set(files_with_summary))

        return files_without_summary

    # ================================================
    # Dropbox management
    # ================================================

    def list_available_files(self) -> list[str]:
        """Return a list of files currently in the directory."""
        if not self.dropbox_dir.exists():
            return []

        available_files = []

        for file in self.dropbox_dir.iterdir():
            if file.is_file() and file.name != "file_metadata.json":
                available_files.append(file.name)

        return available_files

    def store_file_in_dropbox(self, content: str, file_path: Path) -> str:
        """Store file contents in the dropbox."""
        # Add line to manage same filename error
        try:
            self.dropbox_dir.mkdir(parents=True, exist_ok=True)
            path = self.dropbox_dir / file_path.name
            path.write_text(content, encoding="utf-8")

            return ""

        except Exception as e:
            return f"Failed to cache context in workspace storage: {e}"

    def clear_session_dropbox(self) -> str:
        """Remove session dropbox directory when cleaning history."""
        if self.dropbox_dir.exists():
            for child in self.dropbox_dir.iterdir():
                if child.is_file():
                    child.unlink()

            self.dropbox_dir.rmdir()
            return f"Cleared all files in '{self.dropbox_dir}'."

        return f"Error: '{self.dropbox_dir}' not found."


    # ================================================
    # File contents
    # ================================================

    def load_file_content(self, file_path: Path) -> tuple[str, bool]:
        """Return file content as a string using mapped parsers."""
        ext = file_path.suffix.lower()

        if not file_path.exists():
            return f"Error: File {file_path} not found.", False

        # Route to correct parser, fallback to plain text if unknown
        parser = self.formats.get(ext, self._read_txt)

        return parser(file_path), True

    def _load_contents_from_file_list(self, file_paths: list[Path]) -> str:
        """From a list of filenames, returns contents in path as a string."""
        blocks = []

        for path in file_paths:

            filename = path.name
            content, path_exists = self.load_file_content(path)

            if not path_exists:
                print(f"Warning: Error reading file {path}, file not found or unreadable, skipping.")
                continue

            # Format individual file block
            block = (
                "Context from file:\n"
                f"Filename = {filename}\n"
                f"{content}\n"
                f"{'=' * 40}"
            )
            blocks.append(block)

        if not blocks:
            return "No valid file context found."

        return "\n\n".join(blocks)

    # ==================================================
    # Model integration
    # ==================================================

    def _generate_short_summary(self, file_path: Path) -> str:
        """Model generates a short summary about the file content."""
        filename = file_path.name
        content, _ = self.load_file_content(file_path)

        formatted_file_content = (
            f"Context from file:\n"
            f"Filename = {filename}\n"
            f"{content}\n"
        )

        response = LLM.response_with_new_sys_prompt_and_context(
            model=self.model,
            system_prompt=self.gen_summary_prompt,
            prompt=formatted_file_content
        )

        summary = response.message.content

        return summary

    def _add_metadata_and_summary(self, file_paths: list[Path]):
        """Add summary and metadata to files."""
        for path in file_paths:
            filename = path.name
            summary = self._generate_short_summary(path)
            self._add_file_metadata(filename, summary)

    def _structured_file_string(self, filename: str, summary: str) -> str:
        """Format contents for model to read."""
        string = (
            f"Filename = {filename}\n"
            f"Summary = {summary}\n"
            f"{'=' * 40}\n"
        )
        return string

    def require_file_or_not(self, context: list[dict], prompt: str) -> bool:
        """Query model to decide if file context is needed, return 'True' or 'False' only."""
        response = LLM.response_with_new_sys_prompt_and_context(
            model=self.model,
            system_prompt=self.file_or_not_prompt,
            prompt=prompt,
            context=context
        )

        output = response.message.content

        if 'true' in output.lower():
            return True
        else:
            return False

    def get_filenames(self, context: list[dict], available_files_prompt: str) -> tuple[list[Path], list[Path]]:
        """
        Ask model to choose from available in current session to get relevant context,
        formats model output to get clean lists of data.
        """
        response = LLM.response_with_new_sys_prompt_and_context(
            model=self.model,
            system_prompt=self.file_list_prompt,
            prompt=available_files_prompt,
            context=context
        )

        list_of_files = response.message.content

        content                 = response.message.content
        found_file_paths        = []
        not_found_file_paths    = []

        # Reformats generated string to a clean list
        parsed_names = []
        for token in re.split(r"[,\n]", content):
            name = re.sub(r"[*'\"\'\-]", "", token).strip()
            if name:
                parsed_names.append(name)

        # Segregate file paths into found and not found
        for file in parsed_names:
            file_path = self.dropbox_dir / file

            if file_path.exists():
                found_file_paths.append(file)
            else:
                not_found_file_paths.append(file)

        return found_file_paths, not_found_file_paths


    def _file_context(self) -> str:
        """
        Steps:
        1. Add summary and metadata that are not in 'file_metadata.json'.
        2. List all filenames in dropbox.
        3. Format filenames and summary to a string for model to read.
        """
        # Add summary and metadata if not exists
        files_without_summary = self._files_not_in_file_metadata()
        file_paths = [self.dropbox_dir / file for file in files_without_summary]
        self._add_metadata_and_summary(file_paths)

        # Re-sync metadata
        if hasattr(self, "_load_file_metadata"):
            self.file_metadata = self._load_file_metadata() or {}

        # List all available files
        filenames = self.list_available_files()

        # Filenames with its corresponding summary
        blocks = []
        for file in filenames:
            metadata = self.file_metadata.get(file, {})
            summary = metadata.get("summary", "No summary available")

            blocks.append(self._structured_file_string(file, summary))
        
        entries = "\n\n".join(blocks)

        return f"All available files:\n\n{'=' * 40}\n{entries}"

    # ==================================================
    # Auto function
    # ==================================================

    def toggle_auto_read_dropbox(
            self,
            max_tokens: int,
            auto_read_dropbox: bool,
            messages: list[dict],
            prompt: str
    ) -> tuple[str, list[Path], bool]:
        """
        Auto fetches previous file contents ability, return relevant files if its toggled on.

        Model decides from {memory_entries} + {prompt} --> {file_content} from dropbox
        """
        if auto_read_dropbox == True and max_tokens > config.AUTO_READ_DROPBOX_TOKENS:

            # Return nothing if dropbox is empty
            if _is_dir_empty(self.dropbox_dir):
                return "", [], False
            
            require_file = self.require_file_or_not(messages, prompt)

            if require_file == True:

                available_files_prompt = self._file_context()

                found_files, not_found_files = self.get_filenames(messages, available_files_prompt)

                # turn found_files to paths
                found_file_paths = [self.dropbox_dir / file for file in found_files]
                return self._load_contents_from_file_list(found_file_paths), found_files, True

            return "", [], False

        else:
            return "", [], False


    # ==================================================
    # Manual read files with given paths
    # ==================================================

    def read_files_with_context_prompt(
            self,
            context: list[dict],
            file_paths: list[Path] | None,
    ) -> str:
        """Read all files contents from a list of filenames."""
        labelled_content = ["Here are the required context:\n"]

        if file_paths:

            for path in file_paths:

                filename = path.name
                file_content, path_exists = self.load_file_content(path)

                if not path_exists:
                    return ""

                block = (
                    f"Context from file:"
                    f"Filename = {filename}\n"
                    f"{file_content}\n"
                    f"{'=' * 40}"
                )
                labelled_content.append(block)

        else:
            return ""

        return "\n\n".join(labelled_content)
